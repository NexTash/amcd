# Copyright (c) 2025, Umer and contributors
# For license information, please see license.txt

import frappe
import os
from openpyxl import load_workbook
from frappe.utils.file_manager import get_file_path
from frappe.model.document import Document
from frappe.utils.file_manager import save_file
from openpyxl import Workbook


class Assessment(Document):
	def validate(self):
		if self.semester and self.section:
			if not self.assesment_result:
				self.assesment_result = []
				student_doc=frappe.get_all("Student", filters={"semester": self.semester, "section": self.section}, fields=["name","student_name"])
				for doc in student_doc:
					self.append("assesment_result", {
						"roll_no": doc.name,
						"student_name": doc.student_name
					})

	def on_submit(self):
		for result_row in self.assesment_result:
			for clo_row in self.no_clo:
				is_achieved = 1 if result_row.marks and result_row.marks >= clo_row.attainment_level else 0
				status = "CLO Achieved" if is_achieved else "CLO Not Achieved"
				log = frappe.get_doc({
					"doctype": "Student Result Log",
					"student_result": self.name,
					"roll_number": result_row.roll_no,
					"student": result_row.student_name,
					"clo": clo_row.clo,
					"attainment_level": clo_row.attainment_level,
					"obtain_marks": result_row.marks or 0,
					"is_achieved": is_achieved,
					"clo_status": status,
					"section": self.section,
					"semester": self.semester,
				})
				log.insert()


	@frappe.whitelist()
	def download_assessment_results(self):

		# Create Excel workbook and worksheet
		wb = Workbook()
		ws = wb.active
		ws.title = "Assessment Results"

		# Add headers
		ws.append(["Roll No", "Student Name", "Marks"])

		# Add data rows
		for row in self.assesment_result:
			ws.append([
				row.roll_no,
				row.student_name,
				row.marks if row.marks is not None else ""
			])

		# Save to temporary file path
		file_path = f"/tmp/Assessment_Result_{self.name}.xlsx"
		wb.save(file_path)

		# Read and attach file in Frappe
		file_name = f"Assessment_Result_{self.name}.xlsx"
		with open(file_path, "rb") as f:
			filedata = f.read()

		new_file = save_file(file_name, filedata, self.doctype, self.name, is_private=0)

		os.remove(file_path)

		return new_file



	@frappe.whitelist()
	def upload_assessment_results(self):

		file_path = get_file_path(self.attach_file)
		wb = load_workbook(filename=file_path)
		ws = wb.active

		self.set("assesment_result", [])

		for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
			roll_no, student_name, marks = row

			if not roll_no:
				frappe.throw(f"Row {i}: Roll No is required")

			self.append("assesment_result", {
				"roll_no": roll_no,
				"student_name": student_name or "",
				"marks": int(marks) if marks is not None else 0
			})

		self.save()