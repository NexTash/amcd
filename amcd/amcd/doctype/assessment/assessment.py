# Copyright (c) 2025, Umer and contributors
# For license information, please see license.txt

import frappe
import os
from openpyxl import load_workbook
from frappe.utils.file_manager import get_file_path
from frappe.model.document import Document
from frappe.utils.file_manager import save_file
from openpyxl import Workbook


class Assessment(Document):
	@frappe.whitelist()
	def download_assessment_results(self):

		# Create Excel workbook and worksheet
		wb = Workbook()
		ws = wb.active
		ws.title = "Assessment Results"

		# Add headers
		ws.append(["Roll No", "Student Name", "Marks"])

		# Add data rows
		for row in self.assesment_result:
			ws.append([
				row.roll_no,
				row.student_name,
				row.marks if row.marks is not None else ""
			])

		# Save to temporary file path
		file_path = f"/tmp/Assessment_Result_{self.name}.xlsx"
		wb.save(file_path)

		# Read and attach file in Frappe
		file_name = f"Assessment_Result_{self.name}.xlsx"
		with open(file_path, "rb") as f:
			filedata = f.read()

		new_file = save_file(file_name, filedata, self.doctype, self.name, is_private=0)

		os.remove(file_path)

		return new_file



	@frappe.whitelist()
	def upload_assessment_results(self):

		file_path = get_file_path(self.attach_file)
		wb = load_workbook(filename=file_path)
		ws = wb.active

		self.set("assesment_result", [])

		for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
			roll_no, student_name, marks = row

			if not roll_no:
				frappe.throw(f"Row {i}: Roll No is required")

			self.append("assesment_result", {
				"roll_no": roll_no,
				"student_name": student_name or "",
				"marks": int(marks) if marks is not None else 0
			})

		self.save()